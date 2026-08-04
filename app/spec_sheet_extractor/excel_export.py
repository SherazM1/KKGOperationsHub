"""Excel export helpers for the Spec Sheet Extractor module."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from app.spec_sheet_extractor.models import PdfHeaderExtractionResult


DEFAULT_TEMPLATE_PATH = Path("data/sample/Spec_Sheet_Template.xlsx")
TEMPLATE_WORKSHEET_NAME = "Extracted Specs"

EXPECTED_TEMPLATE_HEADERS: tuple[str, ...] = (
    "Customer",
    "Design",
    "rev.",
    "Part",
    "Oppty/Project #",
    "Pieces per set",
    "Board",
    "Corr direction",
    "View",
    "Production Mngr",
    "Designer",
    "ID",
    "Area",
    "Blank width",
    "Blank height",
    "Inches of rule",
    "Date",
    "Upper Special Text",
    "Lower Special Text",
)


class SpecSheetExcelExportError(ValueError):
    """Controlled error for Spec Sheet Excel export failures."""


def _result_values(result: PdfHeaderExtractionResult) -> tuple[str, ...]:
    return (
        result.customer,
        result.design,
        result.revision,
        result.part,
        result.opportunity_project_number,
        result.pieces_per_set,
        result.board,
        result.corr_direction,
        result.view,
        result.production_project_manager,
        result.designer,
        result.id,
        result.area,
        result.blank_width,
        result.blank_height,
        result.inches_of_rule,
        result.date,
        result.upper_special_text,
        result.lower_special_text,
    )


def _validate_template_headers(worksheet: object) -> None:
    headers = tuple(
        worksheet.cell(row=1, column=column).value
        for column in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1)
    )
    if headers != EXPECTED_TEMPLATE_HEADERS:
        raise SpecSheetExcelExportError(
            "Spec Sheet template headers do not match the expected schema."
        )


def _copy_row_style(worksheet: object, source_row: int, target_row: int) -> None:
    for column in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1):
        source_cell = worksheet.cell(row=source_row, column=column)
        target_cell = worksheet.cell(row=target_row, column=column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)


def _prepare_data_rows(worksheet: object, row_count: int) -> None:
    template_style_row = 2 if worksheet.max_row >= 2 else 1
    for row in range(2, max(worksheet.max_row, row_count + 1) + 1):
        _copy_row_style(worksheet, template_style_row, row)
        for column in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1):
            worksheet.cell(row=row, column=column).value = None


def build_spec_sheet_excel(
    extraction_results: Sequence[PdfHeaderExtractionResult],
    *,
    template_path: Path = DEFAULT_TEMPLATE_PATH,
    worksheet_name: str = TEMPLATE_WORKSHEET_NAME,
) -> bytes:
    """Populate the Spec Sheet template and return workbook bytes."""
    if not template_path.exists():
        raise SpecSheetExcelExportError(f"Template file was not found: {template_path}")

    try:
        workbook = load_workbook(template_path)
    except Exception as exc:
        raise SpecSheetExcelExportError(f"Template could not be read: {exc}") from exc

    if worksheet_name not in workbook.sheetnames:
        raise SpecSheetExcelExportError(
            f"Template worksheet was not found: {worksheet_name}"
        )

    worksheet = workbook[worksheet_name]
    _validate_template_headers(worksheet)

    exportable_results = [
        result
        for result in extraction_results
        if result.extraction_status != "Failed extraction"
    ]
    _prepare_data_rows(worksheet, len(exportable_results))

    for row_offset, result in enumerate(exportable_results, start=2):
        for column, value in enumerate(_result_values(result), start=1):
            cell = worksheet.cell(row=row_offset, column=column)
            cell.value = value if value != "" else None
            cell.number_format = "@"
            if column in (18, 19):
                cell.alignment = copy(cell.alignment) if cell.alignment else Alignment()
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent,
                )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
