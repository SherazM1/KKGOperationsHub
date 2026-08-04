"""Tests for Spec Sheet Extractor PDF inventory helpers."""

from __future__ import annotations

from dataclasses import dataclass
import hashlib
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from app.spec_sheet_extractor.excel_export import (
    DEFAULT_TEMPLATE_PATH,
    EXPECTED_TEMPLATE_HEADERS,
    SpecSheetExcelExportError,
    TEMPLATE_WORKSHEET_NAME,
    build_spec_sheet_excel,
)
from app.spec_sheet_extractor import extractor
from app.spec_sheet_extractor.extractor import extract_header_fields_from_uploads, inventory_pdf_uploads
from app.spec_sheet_extractor.models import PdfHeaderExtractionResult, SPEC_SHEET_FIXED_FIELDS
from app.spec_sheet_extractor.zones import HEADER_FIELD_ZONES


@dataclass(frozen=True)
class UploadedFileStub:
    name: str
    data: bytes

    def getvalue(self) -> bytes:
        return self.data


def _pdf_bytes(page_count: int) -> bytes:
    buffer = BytesIO()
    canv = canvas.Canvas(buffer, pagesize=letter)
    for page_number in range(1, page_count + 1):
        canv.drawString(72, 720, f"Page {page_number}")
        canv.showPage()
    canv.save()
    return buffer.getvalue()


def _representative_values() -> dict[str, str]:
    return {
        "Customer": "Fresh Step",
        "Design": "KK260127-02",
        "Revision": "0",
        "Part": "Graphic Box",
        "Opportunity/Project #": "3356",
        "Pieces per set": "2",
        "Board": "200 B",
        "Corr direction": "Vertical",
        "View": "outside",
        "Production/Project Manager": "Marta Espina",
        "Designer": "Andy Bales",
        "ID": "FS-01",
        "Area": "1766.81",
        "Blank width": "53+9/16",
        "Blank height": "38+3/4",
        "Inches of rule": "522+25/32",
        "Date": "07/20/2026",
    }


def _result(
    *,
    index: int = 0,
    page_number: int = 1,
    status: str = "Extracted",
    **overrides: str,
) -> PdfHeaderExtractionResult:
    values = _representative_values()
    values.update(overrides)
    return PdfHeaderExtractionResult(
        source_filename=f"source-{index}.pdf",
        source_file_index=index,
        page_number=page_number,
        extraction_status=status,
        error_message=None if status != "Failed extraction" else "failed",
        customer=values["Customer"],
        design=values["Design"],
        revision=values["Revision"],
        part=values["Part"],
        opportunity_project_number=values["Opportunity/Project #"],
        pieces_per_set=values["Pieces per set"],
        board=values["Board"],
        corr_direction=values["Corr direction"],
        view=values["View"],
        production_project_manager=values["Production/Project Manager"],
        designer=values["Designer"],
        id=values["ID"],
        area=values["Area"],
        blank_width=values["Blank width"],
        blank_height=values["Blank height"],
        inches_of_rule=values["Inches of rule"],
        date=values["Date"],
        upper_special_text=overrides.get("upper_special_text", ""),
        lower_special_text=overrides.get("lower_special_text", ""),
    )


def _workbook_from_export(results: list[PdfHeaderExtractionResult]):
    return load_workbook(BytesIO(build_spec_sheet_excel(results)))


def _label_for_field(field_name: str, manager_label: str) -> str:
    if field_name == "Production/Project Manager":
        return manager_label
    return field_name


def _header_pdf_bytes(
    pages: list[dict[str, str]],
    *,
    manager_label: str = "Production Mngr",
    rotate_degrees: int | None = None,
    ctm_offset: tuple[float, float] | None = None,
) -> bytes:
    buffer = BytesIO()
    canv = canvas.Canvas(buffer, pagesize=letter)
    page_width, page_height = letter
    zone_by_field = {zone.field_name: zone for zone in HEADER_FIELD_ZONES}

    for page_values in pages:
        canv.setFont("Helvetica", 9)
        if ctm_offset is not None:
            canv.saveState()
            canv.translate(*ctm_offset)
        for field_name in SPEC_SHEET_FIXED_FIELDS:
            zone = zone_by_field[field_name]
            x = zone.left * page_width + 4
            y = ((zone.bottom + zone.top) / 2) * page_height
            if ctm_offset is not None:
                x -= ctm_offset[0]
                y -= ctm_offset[1]
            label = _label_for_field(field_name, manager_label)
            value = page_values.get(field_name, "")
            canv.drawString(x, y, f"{label}: {value}")
        if ctm_offset is not None:
            canv.restoreState()
        canv.setFont("Helvetica", 11)
        canv.drawString(72, 420, "BODY DIELINE MEASUREMENT Blank width: 999+1/2")
        canv.drawString(72, 400, "Inches of rule: 999 Date: 01/01/1900")
        canv.showPage()

    canv.save()
    pdf_bytes = buffer.getvalue()
    if rotate_degrees is None:
        return pdf_bytes

    reader = PdfReader(BytesIO(pdf_bytes))
    writer = PdfWriter()
    for page in reader.pages:
        page.rotate(rotate_degrees)
        writer.add_page(page)
    rotated = BytesIO()
    writer.write(rotated)
    return rotated.getvalue()


def _rotated_storage_header_pdf_bytes(pages: list[dict[str, str]]) -> bytes:
    buffer = BytesIO()
    canv = canvas.Canvas(buffer, pagesize=letter)
    raw_width, raw_height = letter
    display_width, display_height = raw_height, raw_width
    zone_by_field = {zone.field_name: zone for zone in HEADER_FIELD_ZONES}

    for page_values in pages:
        canv.setFont("Helvetica", 9)
        for field_name in SPEC_SHEET_FIXED_FIELDS:
            zone = zone_by_field[field_name]
            display_x = zone.left * display_width + 4
            display_y = ((zone.bottom + zone.top) / 2) * display_height
            raw_x = raw_width - display_y
            raw_y = display_x
            value = page_values.get(field_name, "")
            canv.drawString(raw_x, raw_y, f"{field_name}: {value}")
        canv.showPage()
    canv.save()

    reader = PdfReader(BytesIO(buffer.getvalue()))
    writer = PdfWriter()
    for page in reader.pages:
        page.rotate(90)
        writer.add_page(page)
    rotated = BytesIO()
    writer.write(rotated)
    return rotated.getvalue()


def _special_text_pdf_bytes(
    *,
    upper_lines: list[str] | None = None,
    lower_lines: list[str] | None = None,
    extra_lines: list[tuple[float, float, str]] | None = None,
) -> bytes:
    buffer = BytesIO()
    canv = canvas.Canvas(buffer, pagesize=letter)
    page_width, page_height = letter
    values = _representative_values()
    zone_by_field = {zone.field_name: zone for zone in HEADER_FIELD_ZONES}

    canv.setFont("Helvetica", 9)
    for field_name in SPEC_SHEET_FIXED_FIELDS:
        zone = zone_by_field[field_name]
        x = zone.left * page_width + 4
        y = ((zone.bottom + zone.top) / 2) * page_height
        canv.drawString(x, y, f"{field_name}: {values[field_name]}")

    canv.setFont("Helvetica", 10)
    for index, line in enumerate(upper_lines or []):
        canv.drawString(72, 500 - index * 14, line)
    for index, line in enumerate(lower_lines or []):
        canv.drawString(72, 112 - index * 14, line)
    for x, y, line in extra_lines or []:
        canv.drawString(x, y, line)
    canv.save()
    return buffer.getvalue()


def test_inventories_one_valid_single_page_pdf() -> None:
    file_results, page_inventory = inventory_pdf_uploads(
        [UploadedFileStub("single.pdf", _pdf_bytes(1))]
    )

    assert len(file_results) == 1
    assert file_results[0].source_filename == "single.pdf"
    assert file_results[0].page_count == 1
    assert file_results[0].status == "Ready"
    assert len(page_inventory) == 1
    assert page_inventory[0].source_filename == "single.pdf"
    assert page_inventory[0].page_number == 1


def test_inventories_one_valid_multi_page_pdf() -> None:
    file_results, page_inventory = inventory_pdf_uploads(
        [UploadedFileStub("multi.pdf", _pdf_bytes(3))]
    )

    assert file_results[0].page_count == 3
    assert [record.page_number for record in page_inventory] == [1, 2, 3]


def test_multiple_pdfs_preserve_file_and_page_order() -> None:
    file_results, page_inventory = inventory_pdf_uploads(
        [
            UploadedFileStub("a.pdf", _pdf_bytes(2)),
            UploadedFileStub("b.pdf", _pdf_bytes(3)),
        ]
    )

    assert [result.source_filename for result in file_results] == ["a.pdf", "b.pdf"]
    assert [result.source_file_index for result in file_results] == [0, 1]
    assert [
        (record.source_filename, record.source_file_index, record.page_number)
        for record in page_inventory
    ] == [
        ("a.pdf", 0, 1),
        ("a.pdf", 0, 2),
        ("b.pdf", 1, 1),
        ("b.pdf", 1, 2),
        ("b.pdf", 1, 3),
    ]


def test_invalid_pdf_mixed_with_valid_pdf_records_failure_and_valid_pages() -> None:
    file_results, page_inventory = inventory_pdf_uploads(
        [
            UploadedFileStub("bad.pdf", b"not a pdf"),
            UploadedFileStub("good.pdf", _pdf_bytes(2)),
        ]
    )

    assert file_results[0].source_filename == "bad.pdf"
    assert file_results[0].status == "Failed"
    assert file_results[0].page_count == 0
    assert file_results[0].error_message
    assert file_results[1].source_filename == "good.pdf"
    assert file_results[1].status == "Ready"
    assert [record.source_filename for record in page_inventory] == ["good.pdf", "good.pdf"]


def test_page_numbering_starts_at_one() -> None:
    _, page_inventory = inventory_pdf_uploads([UploadedFileStub("pages.pdf", _pdf_bytes(4))])

    assert page_inventory[0].page_number == 1
    assert [record.page_number for record in page_inventory] == [1, 2, 3, 4]


def test_failed_files_do_not_prevent_valid_files_from_being_inventoried() -> None:
    file_results, page_inventory = inventory_pdf_uploads(
        [
            UploadedFileStub("first.pdf", _pdf_bytes(1)),
            UploadedFileStub("bad.pdf", b"%PDF broken"),
            UploadedFileStub("last.pdf", _pdf_bytes(1)),
        ]
    )

    assert [result.status for result in file_results] == ["Ready", "Failed", "Ready"]
    assert [
        (record.source_filename, record.page_number)
        for record in page_inventory
    ] == [("first.pdf", 1), ("last.pdf", 1)]


def test_extracts_all_17_fields_from_representative_page() -> None:
    values = _representative_values()

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("fresh-step.pdf", _header_pdf_bytes([values]))]
    )

    assert len(results) == 1
    row = results[0].to_preview_row()
    assert results[0].extraction_status == "Extracted"
    for field_name, expected_value in values.items():
        assert row[field_name] == expected_value


def test_multi_page_extraction_preserves_order() -> None:
    first = {**_representative_values(), "Design": "KK260002-01"}
    second = {**_representative_values(), "Design": "KK260002-02", "Customer": "Energizer"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("multi-header.pdf", _header_pdf_bytes([first, second]))]
    )

    assert [(result.source_filename, result.page_number) for result in results] == [
        ("multi-header.pdf", 1),
        ("multi-header.pdf", 2),
    ]
    assert [result.design for result in results] == ["KK260002-01", "KK260002-02"]
    assert [result.customer for result in results] == ["Fresh Step", "Energizer"]


def test_rotation_normalization_uses_same_header_zones() -> None:
    values = {**_representative_values(), "Customer": "Rotated Customer", "Date": "02/20/2026"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("rotated.pdf", _rotated_storage_header_pdf_bytes([values]))]
    )

    assert results[0].extraction_status == "Extracted"
    assert results[0].customer == "Rotated Customer"
    assert results[0].blank_width == "53+9/16"
    assert results[0].date == "02/20/2026"


def test_real_coordinate_transformations_use_current_and_text_matrices() -> None:
    values = {**_representative_values(), "Customer": "Transformed Customer"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("transformed.pdf", _header_pdf_bytes([values], ctm_offset=(48, -22)))]
    )

    assert results[0].extraction_status == "Extracted"
    assert results[0].customer == "Transformed Customer"
    assert results[0].opportunity_project_number == "3356"


def test_label_aware_fallback_parses_header_when_coordinate_zones_are_blank(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    values = {
        **_representative_values(),
        "Opportunity/Project #": "Oppty-3075",
        "Pieces per set": "1 Set Required",
    }
    monkeypatch.setattr(extractor, "_collect_zone_text", lambda _page, _zone: "")

    results = extractor.extract_header_fields_from_uploads(
        [UploadedFileStub("fallback.pdf", _header_pdf_bytes([values], manager_label="Project Mngr"))]
    )

    assert results[0].extraction_status == "Extracted"
    assert results[0].customer == "Fresh Step"
    assert results[0].opportunity_project_number == "Oppty-3075"
    assert results[0].pieces_per_set == "1 Set Required"
    assert results[0].production_project_manager == "Marta Espina"


def test_blank_id_field_remains_blank() -> None:
    values = {**_representative_values(), "ID": ""}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("blank-id.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].id == ""
    assert results[0].extraction_status == "Extracted"


def test_zero_field_extraction_becomes_failed_extraction() -> None:
    results = extract_header_fields_from_uploads(
        [UploadedFileStub("no-header.pdf", _pdf_bytes(1))]
    )

    assert results[0].extraction_status == "Failed extraction"
    assert results[0].error_message == "No fixed header fields could be extracted."


def test_partial_extraction_with_legitimate_blanks_is_extracted() -> None:
    values = {field_name: "" for field_name in SPEC_SHEET_FIXED_FIELDS}
    values["Customer"] = "Only Customer"

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("partial.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].extraction_status == "Extracted"
    assert results[0].customer == "Only Customer"
    assert results[0].design == ""


def test_date_returns_strict_mm_dd_yyyy_value() -> None:
    values = {**_representative_values(), "Date": "02/20/2026"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("date.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].date == "02/20/2026"


def test_date_does_not_include_nearby_dimensions() -> None:
    values = {**_representative_values(), "Date": "02/20/2026 43 12 1809.9"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("date-dimensions.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].date == "02/20/2026"


def test_date_remains_blank_when_no_valid_date_exists() -> None:
    values = {**_representative_values(), "Date": "not available"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("no-date.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].date == ""
    assert results[0].extraction_status == "Extracted"


def test_inches_of_rule_removes_trailing_date_label() -> None:
    values = {**_representative_values(), "Inches of rule": "522+25/32 Date:"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("rule-date-label.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].inches_of_rule == "522+25/32"


def test_populated_inaccurately_extracted_field_can_create_warning(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fields = _representative_values()

    def warned_page(_page: object) -> extractor.HeaderFieldExtraction:
        return extractor.HeaderFieldExtraction(
            fields,
            ("Date was corrected by strict date parsing.",),
        )

    monkeypatch.setattr(extractor, "_extract_page_header_fields", warned_page)

    results = extractor.extract_header_fields_from_uploads(
        [UploadedFileStub("warning.pdf", _header_pdf_bytes([fields]))]
    )

    assert results[0].extraction_status == "Extracted with blanks"
    assert results[0].error_message == "Date was corrected by strict date parsing."


def test_fractional_values_are_preserved_exactly() -> None:
    values = {
        **_representative_values(),
        "Blank width": "53+9/16",
        "Blank height": "38+3/4",
        "Inches of rule": "522+25/32",
    }

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("fractions.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].blank_width == "53+9/16"
    assert results[0].blank_height == "38+3/4"
    assert results[0].inches_of_rule == "522+25/32"


def test_opportunity_project_prefix_is_preserved() -> None:
    values = {**_representative_values(), "Opportunity/Project #": "Oppty-3075"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("oppty.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].opportunity_project_number == "Oppty-3075"


def test_pieces_per_set_preserves_wording() -> None:
    values = {**_representative_values(), "Pieces per set": "1 Set Required"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("pieces.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].pieces_per_set == "1 Set Required"


@pytest.mark.parametrize("manager_label", ["Production Mngr", "Project Mngr"])
def test_manager_label_variants_map_to_single_output_field(manager_label: str) -> None:
    values = {**_representative_values(), "Production/Project Manager": "Marta Espina"}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("manager.pdf", _header_pdf_bytes([values], manager_label=manager_label))]
    )

    assert results[0].production_project_manager == "Marta Espina"


def test_one_failed_page_does_not_block_remaining_pages(monkeypatch: pytest.MonkeyPatch) -> None:
    values = _representative_values()
    original_extract_page = extractor._extract_page_header_fields
    call_count = 0

    def fail_second_page(page: object) -> extractor.HeaderFieldExtraction:
        nonlocal call_count
        call_count += 1
        if call_count == 2:
            raise ValueError("page could not be processed")
        return original_extract_page(page)

    monkeypatch.setattr(extractor, "_extract_page_header_fields", fail_second_page)

    results = extractor.extract_header_fields_from_uploads(
        [UploadedFileStub("one-bad-page.pdf", _header_pdf_bytes([values, values, values]))]
    )

    assert [result.extraction_status for result in results] == [
        "Extracted",
        "Failed extraction",
        "Extracted",
    ]
    assert results[1].error_message == "page could not be processed"
    assert results[2].customer == "Fresh Step"


def test_body_dieline_measurement_text_does_not_leak_into_header_fields() -> None:
    values = {**_representative_values(), "Blank width": "", "Inches of rule": ""}

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("body-text.pdf", _header_pdf_bytes([values]))]
    )

    assert results[0].blank_width == ""
    assert results[0].inches_of_rule == ""


def test_upper_special_text_extraction() -> None:
    results = extract_header_fields_from_uploads(
        [UploadedFileStub("upper.pdf", _special_text_pdf_bytes(upper_lines=["4 color - Litho or Digital"]))]
    )

    assert results[0].upper_special_text == "4 color - Litho or Digital"
    assert results[0].lower_special_text == ""


def test_lower_special_text_extraction() -> None:
    results = extract_header_fields_from_uploads(
        [UploadedFileStub("lower.pdf", _special_text_pdf_bytes(lower_lines=["Graphic Box - 1 Shown, 2 Required"]))]
    )

    assert results[0].upper_special_text == ""
    assert results[0].lower_special_text == "Graphic Box - 1 Shown, 2 Required"


def test_multiline_lower_special_text_preserves_line_order() -> None:
    lines = [
        "2025 Holiday Battery Power Stations - 25-ENER-02169",
        "Product Stop",
        "1 Shown - 1 per half Pallet",
        "Outside View - CAD# KK250024-22B",
    ]

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("multi-lower.pdf", _special_text_pdf_bytes(lower_lines=lines))]
    )

    assert results[0].lower_special_text == "\n".join(lines)


def test_repeated_part_and_pieces_wording_is_preserved_in_special_text() -> None:
    lines = ["Blade Set - 1 Set Shown, 1 Set Required"]

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("repeated.pdf", _special_text_pdf_bytes(lower_lines=lines))]
    )

    assert results[0].lower_special_text == lines[0]


def test_copyright_and_standalone_copy_are_excluded_from_special_text() -> None:
    results = extract_header_fields_from_uploads(
        [
            UploadedFileStub(
                "noise.pdf",
                _special_text_pdf_bytes(
                    upper_lines=["COPY", "1 Color"],
                    lower_lines=[
                        "©2026 KENDAL KING. Engineering and graphic designs are the sole intellectual property of Kendal King.",
                        "Fold Over and Glue",
                    ],
                ),
            )
        ]
    )

    assert results[0].upper_special_text == "1 Color"
    assert results[0].lower_special_text == "Fold Over and Glue"


def test_dimension_only_fragments_are_excluded_from_special_text() -> None:
    results = extract_header_fields_from_uploads(
        [
            UploadedFileStub(
                "dimensions.pdf",
                _special_text_pdf_bytes(
                    upper_lines=["53+9/16", "1 x 3 x 1/8", "full litho on this piece"],
                ),
            )
        ]
    )

    assert results[0].upper_special_text == "full litho on this piece"


def test_instruction_text_containing_dimensions_is_preserved() -> None:
    instruction = 'Requires (2) 1" x 3" x 1/16" Pieces of Remo 1 Side'

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("instruction-dimensions.pdf", _special_text_pdf_bytes(lower_lines=[instruction]))]
    )

    assert results[0].lower_special_text == instruction


def test_no_fragment_appears_in_both_upper_and_lower_special_text() -> None:
    line = "1 Required Per Display"

    results = extract_header_fields_from_uploads(
        [UploadedFileStub("boundary.pdf", _special_text_pdf_bytes(extra_lines=[(72, 120, line)]))]
    )

    occurrences = results[0].upper_special_text.count(line) + results[0].lower_special_text.count(line)
    assert occurrences == 1


def test_blank_special_text_zones_do_not_create_warnings() -> None:
    results = extract_header_fields_from_uploads(
        [UploadedFileStub("blank-special.pdf", _special_text_pdf_bytes())]
    )

    assert results[0].upper_special_text == ""
    assert results[0].lower_special_text == ""
    assert results[0].extraction_status == "Extracted"
    assert results[0].error_message is None


def test_fixed_header_fields_do_not_leak_into_special_text() -> None:
    results = extract_header_fields_from_uploads(
        [
            UploadedFileStub(
                "header-protection.pdf",
                _special_text_pdf_bytes(
                    upper_lines=["Customer:", "Fresh Step", "Design:", "KK260127-02", "1 Color"]
                ),
            )
        ]
    )

    assert results[0].upper_special_text == "1 Color"


def test_spec_sheet_template_loads_successfully() -> None:
    workbook = load_workbook(DEFAULT_TEMPLATE_PATH)

    assert TEMPLATE_WORKSHEET_NAME in workbook.sheetnames


def test_expected_19_template_headers_are_validated() -> None:
    workbook = load_workbook(DEFAULT_TEMPLATE_PATH)
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    headers = tuple(worksheet.cell(row=1, column=column).value for column in range(1, 20))

    assert headers == EXPECTED_TEMPLATE_HEADERS


def test_header_mismatch_creates_controlled_error(tmp_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = TEMPLATE_WORKSHEET_NAME
    for column, header in enumerate(EXPECTED_TEMPLATE_HEADERS, start=1):
        worksheet.cell(row=1, column=column).value = header
    worksheet.cell(row=1, column=3).value = "wrong header"
    template_path = tmp_path / "bad_template.xlsx"
    workbook.save(template_path)

    with pytest.raises(SpecSheetExcelExportError):
        build_spec_sheet_excel([_result()], template_path=template_path)


def test_one_extraction_result_populates_row_2_correctly() -> None:
    workbook = _workbook_from_export(
        [
            _result(
                Date="02/20/2026",
                **{
                    "Blank width": "53+9/16",
                    "Inches of rule": "522+25/32",
                    "upper_special_text": "4 color - Litho or Digital",
                    "lower_special_text": "Graphic Box - 1 Shown, 2 Required",
                },
            )
        ]
    )
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    assert worksheet.cell(row=2, column=1).value == "Fresh Step"
    assert worksheet.cell(row=2, column=2).value == "KK260127-02"
    assert worksheet.cell(row=2, column=17).value == "02/20/2026"
    assert worksheet.cell(row=2, column=18).value == "4 color - Litho or Digital"
    assert worksheet.cell(row=2, column=19).value == "Graphic Box - 1 Shown, 2 Required"


def test_multiple_export_results_preserve_order() -> None:
    workbook = _workbook_from_export(
        [
            _result(index=0, page_number=1, Customer="First", Design="KK1"),
            _result(index=1, page_number=1, Customer="Second", Design="KK2"),
        ]
    )
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    assert [worksheet.cell(row=row, column=1).value for row in (2, 3)] == ["First", "Second"]
    assert [worksheet.cell(row=row, column=2).value for row in (2, 3)] == ["KK1", "KK2"]


def test_45_page_results_create_45_populated_rows() -> None:
    workbook = _workbook_from_export(
        [_result(index=0, page_number=page, Design=f"KK-{page:02}") for page in range(1, 46)]
    )
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    assert worksheet.cell(row=46, column=2).value == "KK-45"
    assert worksheet.cell(row=47, column=1).value is None


def test_legitimate_blanks_remain_blank_cells() -> None:
    workbook = _workbook_from_export([_result(ID="", upper_special_text="", lower_special_text="")])
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    assert worksheet.cell(row=2, column=12).value is None
    assert worksheet.cell(row=2, column=18).value is None
    assert worksheet.cell(row=2, column=19).value is None


def test_export_preserves_date_fraction_and_multiline_special_text_exactly() -> None:
    workbook = _workbook_from_export(
        [
            _result(
                Date="02/20/2026",
                **{
                    "Blank width": "53+9/16",
                    "lower_special_text": "Line One\nLine Two",
                },
            )
        ]
    )
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    assert worksheet.cell(row=2, column=14).value == "53+9/16"
    assert worksheet.cell(row=2, column=17).value == "02/20/2026"
    assert worksheet.cell(row=2, column=19).value == "Line One\nLine Two"


def test_source_filename_page_status_and_error_are_not_exported() -> None:
    workbook = _workbook_from_export([_result()])
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]
    exported_headers = [worksheet.cell(row=1, column=column).value for column in range(1, 20)]

    assert "Source File" not in exported_headers
    assert "Page Number" not in exported_headers
    assert "Status" not in exported_headers
    assert "Error" not in exported_headers


def test_failed_extraction_pages_are_excluded_from_export() -> None:
    workbook = _workbook_from_export(
        [
            _result(Customer="Good"),
            _result(Customer="Bad", status="Failed extraction"),
            _result(Customer="Also Good"),
        ]
    )
    worksheet = workbook[TEMPLATE_WORKSHEET_NAME]

    assert worksheet.cell(row=2, column=1).value == "Good"
    assert worksheet.cell(row=3, column=1).value == "Also Good"
    assert worksheet.cell(row=4, column=1).value is None


def test_source_template_file_is_not_modified() -> None:
    before = hashlib.sha256(DEFAULT_TEMPLATE_PATH.read_bytes()).hexdigest()

    build_spec_sheet_excel([_result()])

    after = hashlib.sha256(DEFAULT_TEMPLATE_PATH.read_bytes()).hexdigest()
    assert after == before


def test_generated_workbook_can_be_reopened_successfully() -> None:
    excel_bytes = build_spec_sheet_excel([_result()])

    workbook = load_workbook(BytesIO(excel_bytes))

    assert TEMPLATE_WORKSHEET_NAME in workbook.sheetnames
